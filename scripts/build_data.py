#!/usr/bin/env python3
"""
Armenia 2026 Legislative Election Atlas — data pipeline.

Single source of truth: Central Electoral Commission (CEC) of Armenia.
  - Results by polling station : https://www.elections.am/File/ElectionResult?electionId=28826
  - Polling station registry   : https://www.elections.am/File/SubDistrictsToExcel?electionId=28826

This script joins the two official workbooks (results keyed by polling station,
registry providing marz / community / constituency) and produces clean, versioned
artefacts at every aggregation level: polling station -> community -> marz -> national.

Each election lives under data/<ELECTION>/ (default 2026; override with the
ELECTION environment variable). Outputs (all derived, reproducible):
  data/<ELECTION>/clean/stations.{parquet,csv}      full granular table (stations x parties)
  data/<ELECTION>/clean/communities.{parquet,csv}   aggregated by community
  data/<ELECTION>/clean/marz.{parquet,csv}          aggregated by marz (11 provinces)
  data/<ELECTION>/national.json                     national party totals + meta (for the site)
  data/<ELECTION>/marz.json                         per-marz totals, winner, margin, shares (for the map)
  data/<ELECTION>/parties.json                      trilingual party metadata + national result + best/worst marz
  data/<ELECTION>/meta.json                         provenance, timestamps, integrity checks
"""
from __future__ import annotations

import datetime as _dt
import json
import os
import pathlib
import re
import sys

import openpyxl
import pandas as pd

ROOT = pathlib.Path(__file__).resolve().parent.parent
ELECTION = os.environ.get("ELECTION", "2026")
RAW = ROOT / "data" / ELECTION / "raw"
CLEAN = ROOT / "data" / ELECTION / "clean"
DATA = ROOT / "data" / ELECTION

RESULTS_XLSX = RAW / "cec_results_by_station.xlsx"
REGISTRY_XLSX = RAW / "cec_subdistricts.xlsx"

# --- Administrative columns in the results workbook (0-indexed) ---
# These are identical across the CEC results workbooks (2017–2026).
COL_TEC = 0           # Territorial Electoral Commission number
COL_STATION = 1       # Polling station code, e.g. "1/1"
COL_REGISTERED = 5    # Total number of voters (registered)
COL_PARTICIPANTS = 6  # Total number of participants (ballots cast)
COL_INVALID = 14      # Invalid ballots

# ---------------------------------------------------------------------------
# Per-election configuration. Ballot (party) columns and metadata differ by
# election; select with the ELECTION environment variable (default 2026).
# Each party tuple: (col, id, type, hy, en, fr, abbr, leader, bloc, family, color)
# ---------------------------------------------------------------------------
ELECTIONS = {
    "2026": {
        "election": "National Assembly of Armenia — ordinary election, 7 June 2026",
        "date": "2026-06-07",
        "election_id": 28826,
        "threshold_party_pct": 4.0,
        "threshold_alliance_pct": 8.0,
        "total_seats": 105,
        "col_inaccuracy": 33,
        "seats": {"civil_contract": 64, "strong_armenia": 29, "armenia_alliance": 12},
        "parties": [
            (29, "civil_contract", "party",
             "«Քաղաքացիական պայմանագիր» կուսակցություն", "Civil Contract", "Contrat civil",
             "ՔՊ / CC", "Nikol Pashinyan", "government", "liberal-reformist", "#F58220"),
            (17, "strong_armenia", "alliance",
             "«Ուժեղ Հայաստան» դաշինք", "Strong Armenia Alliance", "Alliance « Arménie forte »",
             "ՈՒՀ / SA", "Samvel Karapetyan", "opposition", "conservative-sovereigntist", "#C8102E"),
            (30, "armenia_alliance", "alliance",
             "«Հայաստան» դաշինք", "Armenia Alliance", "Alliance « Arménie »",
             "ՀԴ / AA", "Robert Kocharyan", "opposition", "national-conservative", "#1B3A6B"),
            (21, "prosperous_armenia", "party",
             "«Բարգավաճ Հայաստան» կուսակցություն", "Prosperous Armenia", "Arménie prospère",
             "ԲՀԿ / PA", "Gagik Tsarukyan", "below-threshold", "centrist-populist", "#E8A300"),
            (20, "wings_of_unity", "party",
             "«Միասնության թևեր» կուսակցություն", "Wings of Unity", "Ailes de l’unité",
             "ՄԹ / WU", "Mane Tandilyan", "below-threshold", "centrist", "#6A4C93"),
            (18, "meritocratic", "party",
             "Հայաստանի շնորհապետական կուսակցություն", "Meritocratic Party of Armenia", "Parti méritocratique d’Arménie",
             "ՀՇԿ / MP", None, "below-threshold", "technocratic", "#0F8B8D"),
            (28, "democracy_law_discipline", "party",
             "«Դեմոկրատիա, օրենք, կարգապահություն» կուսակցություն", "Democracy, Law and Discipline", "Démocratie, loi et discipline",
             "ԴՕԿ / DLD", None, "below-threshold", "law-and-order", "#7D8597"),
            (19, "new_power", "party",
             "«Նոր ուժ» ռեֆորմիստական կուսակցություն", "New Power Reformist Party", "Parti réformiste « Nouvelle Force »",
             "ՆՈՒ / NP", None, "below-threshold", "reformist", "#2A9D8F"),
            (16, "against_all", "party",
             "«Բոլորին դեմ եմ» ժողովրդավարական կուսակցություն", "“Against All” Democratic Party", "Parti démocratique « Contre tous »",
             "ԲԴԵ / AAll", None, "below-threshold", "protest", "#8D99AE"),
            (25, "hanrapetutyun", "party",
             "«Հանրապետություն» կուսակցություն", "Hanrapetutyun (Republic) Party", "Parti Hanrapetoutioun (République)",
             "Հանր / Rep", "Aram Sargsyan", "below-threshold", "liberal-republican", "#5B8C5A"),
            (32, "bright_armenia", "party",
             "«Լուսավոր Հայաստան» կուսակցություն", "Bright Armenia", "Arménie lumineuse",
             "ԼՀԿ / BA", "Edmon Marukyan", "below-threshold", "liberal", "#4895EF"),
            (31, "for_the_republic", "alliance",
             "«Հանուն Հանրապետության» ժողովրդավարության պաշտպանների դաշինք", "“In the Name of the Republic” Democracy Defenders Alliance",
             "Alliance « Au nom de la République »",
             "ՀՀ / NR", None, "below-threshold", "republican", "#9C6644"),
            (22, "national_democratic_pole", "party",
             "«Ազգային-ժողովրդավարական բևեռ» համահայկական կուսակցություն", "Pan-Armenian National Democratic Pole", "Pôle national-démocrate panarménien",
             "ԱԺԲ / NDP", None, "below-threshold", "national-democratic", "#BC6C25"),
            (27, "democratic_consolidation", "party",
             "«Ժողովրդավարական համախմբում» կուսակցություն", "Democratic Consolidation Party", "Consolidation démocratique",
             "ԺՀ / DC", None, "below-threshold", "centrist", "#A3B18A"),
            (24, "anc", "party",
             "«Հայ ազգային կոնգրես» կուսակցություն", "Armenian National Congress", "Congrès national arménien",
             "ՀԱԿ / ANC", "Levon Ter-Petrosyan", "below-threshold", "social-liberal", "#577590"),
            (26, "christian_democratic", "party",
             "Քրիստոնեա-ժողովրդավարական կուսակցություն", "Christian Democratic Party", "Parti chrétien-démocrate",
             "ՔԺԿ / CD", None, "below-threshold", "christian-democratic", "#B5838D"),
            (23, "kochari", "party",
             "«Քոչարի ազգային վերածնունդ և ազգի զարթոնք» կուսակցություն", "Kochari National Revival and Awakening", "Kochari — Renaissance nationale",
             "Քոչ / Koch", None, "below-threshold", "nationalist", "#6D6875"),
            (15, "reformists", "party",
             "Ռեֆորմիստների կուսակցություն", "Reformists Party", "Parti des réformistes",
             "Ռեֆ / Ref", None, "below-threshold", "reformist", "#B08968"),
        ],
    },
    "2021": {
        "election": "National Assembly of Armenia — early election, 20 June 2021",
        "date": "2021-06-20",
        "election_id": 27697,
        "threshold_party_pct": 5.0,
        "threshold_alliance_pct": 7.0,
        "total_seats": 107,
        "col_inaccuracy": 40,
        # In June 2021 communities were still pre-consolidation (~500). Re-aggregate
        # to the modern consolidated municipalities so the community level matches
        # 2026 and is free of cross-province homonyms.
        "consolidate_from": "2026",
        "seats": {"civil_contract": 71, "armenia_alliance": 29, "i_have_honor": 7},
        "parties": [
            (17, "civil_contract", "party",
             "«Քաղաքացիական պայմանագիր» կուսակցություն", "Civil Contract", "Contrat civil",
             "ՔՊ / CC", "Nikol Pashinyan", "government", "liberal-reformist", "#F58220"),
            (37, "armenia_alliance", "alliance",
             "«Հայաստան» դաշինք", "Armenia Alliance", "Alliance « Arménie »",
             "ՀԴ / AA", "Robert Kocharyan", "opposition", "national-conservative", "#1B3A6B"),
            (20, "i_have_honor", "alliance",
             "«Պատիվ ունեմ» դաշինք", "I Have Honor Alliance", "Alliance « J’ai l’honneur »",
             "ՊՈւ / IHH", "Artur Vanetsyan", "opposition", "national-conservative", "#6A2E8F"),
            (28, "prosperous_armenia", "party",
             "«Բարգավաճ Հայաստան» կուսակցություն", "Prosperous Armenia", "Arménie prospère",
             "ԲՀԿ / PA", "Gagik Tsarukyan", "below-threshold", "centrist-populist", "#E8A300"),
            (23, "bright_armenia", "party",
             "«Լուսավոր Հայաստան» կուսակցություն", "Bright Armenia", "Arménie lumineuse",
             "ԼՀԿ / BA", "Edmon Marukyan", "below-threshold", "liberal", "#4895EF"),
            (16, "anc", "party",
             "«Հայ ազգային կոնգրես» կուսակցություն", "Armenian National Congress", "Congrès national arménien",
             "ՀԱԿ / ANC", "Levon Ter-Petrosyan", "below-threshold", "social-liberal", "#577590"),
            (35, "liberal", "party",
             "Ազատական կուսակցություն", "Liberal Party", "Parti libéral",
             "Ազ / Lib", None, "below-threshold", "liberal", "#4361EE"),
            (25, "hanrapetutyun", "party",
             "«Հանրապետություն» կուսակցություն", "Hanrapetutyun (Republic) Party", "Parti Hanrapetoutioun (République)",
             "Հանր / Rep", "Aram Sargsyan", "below-threshold", "liberal-republican", "#5B8C5A"),
            (38, "national_democratic_pole", "party",
             "«Ազգային-ժողովրդավարական բևեռ» համահայկական կուսակցություն", "Pan-Armenian National Democratic Pole", "Pôle national-démocrate panarménien",
             "ԱԺԲ / NDP", None, "below-threshold", "national-democratic", "#9D4EDD"),
            (32, "shirinyan_babajanyan", "alliance",
             "«Շիրինյան-Բաբաջանյան ժողովրդավարների դաշինք»", "Shirinyan-Babajanyan Alliance of Democrats", "Alliance des démocrates Chirinian-Babadjanian",
             "ՇԲ / SB", None, "below-threshold", "social-democratic", "#7D8597"),
            (30, "movement_5165", "party",
             "«5165 ազգային պահպանողական շարժում» կուսակցություն", "5165 National Conservative Movement", "Mouvement national conservateur 5165",
             "5165 / 5165", None, "below-threshold", "national-conservative", "#6D6875"),
            (33, "national_agenda", "party",
             "«Ազգային օրակարգ» կուսակցություն", "National Agenda Party", "Parti de l’agenda national",
             "ԱՕ / NA", None, "below-threshold", "national-conservative", "#8A5A44"),
            (36, "european", "party",
             "Հայաստանի եվրոպական կուսակցություն", "European Party of Armenia", "Parti européen d’Arménie",
             "ՀԵԿ / EPA", None, "below-threshold", "pro-european", "#3F37C9"),
            (39, "sovereign_armenia", "party",
             "«Ինքնիշխան Հայաստան» կուսակցություն", "Sovereign Armenia Party", "Parti Arménie souveraine",
             "ԻՀ / SovA", None, "below-threshold", "sovereigntist", "#6A4C93"),
            (15, "fair_armenia", "party",
             "«Արդար Հայաստան» կուսակցություն", "Fair Armenia", "Arménie juste",
             "ԱՀ / FA", None, "below-threshold", "centrist", "#B08968"),
            (18, "zartonk", "party",
             "«Զարթոնք» ազգային քրիստոնեական կուսակցություն", "Zartonk National Christian Party", "Parti national-chrétien Zartonk",
             "Զարթ / Zar", None, "below-threshold", "christian-democratic", "#9C6644"),
            (19, "liberty", "party",
             "Ազատություն կուսակցություն", "Liberty Party", "Parti de la liberté",
             "Ազատ / Lty", None, "below-threshold", "national-liberal", "#2A9D8F"),
            (21, "united_homeland", "party",
             "Միասնական հայրենիք կուսակցություն", "United Homeland Party", "Parti de la patrie unie",
             "ՄՀ / UH", None, "below-threshold", "nationalist", "#8D6E63"),
            (22, "pan_armenian_statehood", "party",
             "Համահայկական ազգային պետականություն (ՀԱՊ) կուսակցություն", "Pan-Armenian National Statehood (HAP)", "État national panarménien (HAP)",
             "ՀԱՊ / HAP", None, "below-threshold", "nationalist", "#BC6C25"),
            (24, "our_home_armenia", "party",
             "«Մեր տունը Հայաստանն է» կուսակցություն", "Our Home is Armenia", "Notre maison, c’est l’Arménie",
             "ՄՏՀ / OHA", None, "below-threshold", "centrist", "#6D8B74"),
            (26, "homeland_of_armenians", "party",
             "«Հայոց հայրենիք» կուսակցություն", "Homeland of Armenians Party", "Parti de la patrie des Arméniens",
             "ՀՀ / HoA", None, "below-threshold", "nationalist", "#A4133C"),
            (27, "free_homeland", "alliance",
             "Ազատ հայրենիք դաշինք", "Free Homeland Alliance", "Alliance Patrie libre",
             "ԱՀԴ / FH", None, "below-threshold", "centrist", "#3A5A40"),
            (29, "democratic_party", "party",
             "Հայաստանի դեմոկրատական կուսակցություն", "Democratic Party of Armenia", "Parti démocrate d’Arménie",
             "ՀԴԿ / DPA", None, "below-threshold", "social-democratic", "#468FAF"),
            (31, "citizens_decision", "party",
             "«Քաղաքացու որոշում» սոցիալ-դեմոկրատական կուսակցություն", "Citizen’s Decision Social-Democratic Party", "Parti social-démocrate « Décision citoyenne »",
             "ՔՈ / CD", None, "below-threshold", "social-democratic", "#B5838D"),
            (34, "verelk", "party",
             "«Վերելք» կուսակցություն", "Verelk (Rise) Party", "Parti Verelk (Essor)",
             "Վեր / Ver", None, "below-threshold", "centrist", "#588157"),
        ],
    },
}

if ELECTION not in ELECTIONS:
    sys.exit(f"Unknown ELECTION={ELECTION!r}; known: {', '.join(ELECTIONS)}")
CFG = ELECTIONS[ELECTION]

COL_INACCURACY = CFG["col_inaccuracy"]
PARTIES = CFG["parties"]
SOURCE = {
    "results_by_station": f"https://www.elections.am/File/ElectionResult?electionId={CFG['election_id']}",
    "polling_station_registry": f"https://www.elections.am/File/SubDistrictsToExcel?electionId={CFG['election_id']}",
    "authority": "Central Electoral Commission of the Republic of Armenia (elections.am)",
    "election": CFG["election"],
    "boundaries": "geoBoundaries ARM ADM1 (CC-BY 4.0)",
}

PARTY_BY_COL = {p[0]: p[1] for p in PARTIES}
PARTY_IDS = [p[1] for p in PARTIES]
PARTY_META = {
    p[1]: dict(col=p[0], type=p[2], name_hy=p[3], name_en=p[4], name_fr=p[5],
               abbr=p[6], leader=p[7], bloc=p[8], family=p[9], color=p[10])
    for p in PARTIES
}

# Marz: Armenian label (registry) -> iso / trilingual names. ISO from geoBoundaries.
MARZ = {
    "Երևան":      ("AM-ER", "Yerevan", "Erevan", "Երևան"),
    "Արագածոտն":  ("AM-AG", "Aragatsotn", "Aragatsotn", "Արագածոտն"),
    "Արարատ":     ("AM-AR", "Ararat", "Ararat", "Արարատ"),
    "Արմավիր":    ("AM-AV", "Armavir", "Armavir", "Արմավիր"),
    "Գեղարքունիք": ("AM-GR", "Gegharkunik", "Guégharkounik", "Գեղարքունիք"),
    "Լոռի":       ("AM-LO", "Lori", "Lorri", "Լոռի"),
    "Կոտայք":     ("AM-KT", "Kotayk", "Kotayk", "Կոտայք"),
    "Շիրակ":      ("AM-SH", "Shirak", "Chirak", "Շիրակ"),
    "Սյունիք":    ("AM-SU", "Syunik", "Syunik", "Սյունիք"),
    "Վայոց ձոր":  ("AM-VD", "Vayots Dzor", "Vayots Dzor", "Վայոց ձոր"),
    "Տավուշ":     ("AM-TV", "Tavush", "Tavouch", "Տավուշ"),
}
# Reverse lookup keyed by English name -> (iso, en, fr, hy)
MARZ_INV = {v[1]: v for v in MARZ.values()}

# --- consolidation of pre-reform communities into modern municipalities --------
_CDESIG = re.compile(r"\s+(գյուղ|քաղաք|կայարան)\s*$")


def _cnorm(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s*\([^)]*\)\s*", " ", s)   # drop parenthetical qualifiers
    s = re.sub(r"\s+", " ", s).strip()
    s = _CDESIG.sub("", s)                     # drop "գյուղ"/"քաղաք"/…
    return s.replace(" յ", "յ").replace("եվ", "և")


# Pre-reform communities the source registry resolves differently (province
# changed or community dissolved by the reform). Keyed by (marz_hy, _cnorm).
MANUAL_CONSOLIDATION = {
    ("Շիրակ", "Արփի"): "Ամասիա",        # Arpi community (centre Berdashen) merged into Amasia
    ("Կոտայք", "Նոր Երզնկա"): "Նաիրի",  # Nor Yerznka -> Nairi (centre Yeghvard)
}


def load_consolidation_map(source_election: str) -> dict:
    """(marz_hy, normalized settlement/community) -> consolidated community name,
    derived from the source election's registry (col1 community, col2 settlement)."""
    path = ROOT / "data" / source_election / "raw" / "cec_subdistricts.xlsx"
    if not path.exists():
        sys.exit(f"Consolidation source registry missing: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    m = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 5 or not r[4]:
            continue
        marz, comm, sett = r[0], r[1], r[2]
        if not comm:
            continue
        m[(marz, _cnorm(comm))] = comm
        if sett:
            m[(marz, _cnorm(sett))] = comm
    return m


def _num(v):
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def read_results() -> pd.DataFrame:
    wb = openpyxl.load_workbook(RESULTS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # skip 2 title rows + header
        station = row[COL_STATION]
        if not station or "/" not in str(station):
            continue  # totals / blank rows
        rec = {
            "tec": str(row[COL_TEC]).zfill(2) if row[COL_TEC] else None,
            "station": str(station).strip(),
            "registered": _num(row[COL_REGISTERED]),
            "ballots_cast": _num(row[COL_PARTICIPANTS]),
            "invalid": _num(row[COL_INVALID]),
            "inaccuracy": _num(row[COL_INACCURACY]),
        }
        for col, pid in PARTY_BY_COL.items():
            rec[pid] = _num(row[col])
        records.append(rec)
    df = pd.DataFrame.from_records(records)
    df["valid"] = df[PARTY_IDS].sum(axis=1)
    return df


def read_registry() -> pd.DataFrame:
    wb = openpyxl.load_workbook(REGISTRY_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        marz, community, settlement, constituency, station, address = r[:6]
        if not station:
            continue
        rows.append({
            "station": str(station).strip(),
            "marz_hy": (marz or "").strip(),
            "community_hy": (community or "").strip(),
            "settlement_hy": (settlement or "").strip() or None,
            "constituency": str(constituency).strip() if constituency else None,
            "address": (address or "").strip(),
        })
    return pd.DataFrame.from_records(rows)


def winner_info(shares: dict) -> dict:
    ordered = sorted(shares.items(), key=lambda kv: kv[1]["votes"], reverse=True)
    top = ordered[0]
    second = ordered[1] if len(ordered) > 1 else (None, {"pct": 0})
    return {
        "winner": top[0],
        "winner_pct": top[1]["pct"],
        "margin": round(top[1]["pct"] - second[1]["pct"], 2),
        "runner_up": second[0],
        "order": [pid for pid, _ in ordered],
    }


def aggregate(df: pd.DataFrame, by: list[str]) -> pd.DataFrame:
    agg = {c: "sum" for c in PARTY_IDS}
    agg.update({"registered": "sum", "ballots_cast": "sum",
                "invalid": "sum", "valid": "sum"})
    g = df.groupby(by, as_index=False).agg(agg)
    g["stations"] = df.groupby(by, as_index=False).size()["size"].values
    return g


def shares_dict(row) -> dict:
    valid = max(int(row["valid"]), 1)
    out = {}
    for pid in PARTY_IDS:
        v = int(row[pid])
        out[pid] = {"votes": v, "pct": round(100 * v / valid, 3)}
    return out


def build():
    if not RESULTS_XLSX.exists() or not REGISTRY_XLSX.exists():
        sys.exit(f"Raw CEC workbooks missing in {RAW}/. Run scripts/fetch_source.sh first.")

    CLEAN.mkdir(parents=True, exist_ok=True)
    results = read_results()
    registry = read_registry()

    df = results.merge(registry, on="station", how="left")

    # Settlement/locality label: CEC "Բնակավայր" when present, otherwise the
    # community column (Yerevan districts, or pre-reform community in 2021).
    def _locality(row):
        sett = row["settlement_hy"]
        if pd.notna(sett) and str(sett).strip():
            return str(sett).strip()
        return str(row["community_hy"] or "").strip()

    df["locality_hy"] = df.apply(_locality, axis=1)

    # Re-aggregate pre-reform communities into modern consolidated municipalities
    # (keeps each station in its own province; only the community label changes).
    src = CFG.get("consolidate_from")
    if src:
        cmap = load_consolidation_map(src)
        unresolved = set()

        def _consolidate(row):
            key = (row["marz_hy"], _cnorm(row["community_hy"]))
            hit = MANUAL_CONSOLIDATION.get(key) or cmap.get(key)
            if hit:
                return hit
            unresolved.add(row["community_hy"])
            return row["community_hy"]

        df["community_hy"] = df.apply(_consolidate, axis=1)
        print(f"   consolidated to {df['community_hy'].nunique()} communities "
              f"(source {src}); {len(unresolved)} unresolved", file=sys.stderr)
        if unresolved:
            print("   unresolved:", ", ".join(sorted(unresolved)), file=sys.stderr)

    # Integrity guard against "non-territorial" ballots (electronic votes by
    # diplomats / military, any future out-of-country line) being silently folded
    # into a province. Such rows carry a marz label absent from the 11-province
    # MARZ table; we surface them and keep them out of the geographic aggregation.
    df["marz_iso"] = df["marz_hy"].map(lambda x: MARZ.get(x, (None,))[0])
    df["marz_en"] = df["marz_hy"].map(lambda x: MARZ.get(x, (None, None))[1])
    unmatched = df[df["marz_iso"].isna()]
    out_of_marz_votes = int(unmatched["valid"].sum())
    if len(unmatched):
        labels = sorted({(r or "(blank)") for r in unmatched["marz_hy"].fillna("")})
        print(f"WARNING: {len(unmatched)} stations outside the 11 provinces "
              f"({out_of_marz_votes} valid votes) — kept national-only, not a province. "
              f"Labels: {labels}", file=sys.stderr)

    # --- station level (full granular) ---
    cols = (["tec", "station", "constituency", "marz_hy", "marz_en", "marz_iso",
             "community_hy", "address", "registered", "ballots_cast",
             "invalid", "valid"] + PARTY_IDS + ["inaccuracy"])
    stations = df[cols].copy()
    stations["turnout_pct"] = (100 * stations["ballots_cast"] /
                               stations["registered"].clip(lower=1)).round(2)
    stations.to_parquet(CLEAN / "stations.parquet", index=False)
    stations.to_csv(CLEAN / "stations.csv", index=False)

    # --- community level ---
    comm = aggregate(df, ["marz_en", "marz_iso", "community_hy"])
    comm["turnout_pct"] = (100 * comm["ballots_cast"] /
                           comm["registered"].clip(lower=1)).round(2)
    comm.to_parquet(CLEAN / "communities.parquet", index=False)
    comm.to_csv(CLEAN / "communities.csv", index=False)

    # --- settlement / locality level (within consolidated community) ---
    sett = aggregate(df, ["marz_en", "marz_iso", "community_hy", "locality_hy"])
    sett["turnout_pct"] = (100 * sett["ballots_cast"] /
                           sett["registered"].clip(lower=1)).round(2)
    sett.to_parquet(CLEAN / "settlements.parquet", index=False)
    sett.to_csv(CLEAN / "settlements.csv", index=False)

    # --- marz level ---
    marz = aggregate(df, ["marz_en", "marz_iso", "marz_hy"])
    marz["turnout_pct"] = (100 * marz["ballots_cast"] /
                           marz["registered"].clip(lower=1)).round(2)
    marz.to_parquet(CLEAN / "marz.parquet", index=False)
    marz.to_csv(CLEAN / "marz.csv", index=False)

    # --- national ---
    nat_valid = int(results["valid"].sum())
    nat_cast = int(results["ballots_cast"].sum())
    nat_reg = int(results["registered"].sum())
    nat_invalid = int(results["invalid"].sum())
    national_parties = []
    for pid in PARTY_IDS:
        v = int(results[pid].sum())
        national_parties.append({
            "id": pid, "votes": v,
            "pct": round(100 * v / nat_valid, 4),
            **{k: PARTY_META[pid][k] for k in
               ("type", "name_hy", "name_en", "name_fr", "abbr",
                "leader", "bloc", "family", "color")},
        })
    national_parties.sort(key=lambda p: p["votes"], reverse=True)

    # seats (CEC final allocation, incl. national-minority + stabilising seats)
    seats = CFG["seats"]
    for p in national_parties:
        p["seats"] = seats.get(p["id"], 0)

    national = {
        "election": SOURCE["election"],
        "date": CFG["date"],
        "threshold_party_pct": CFG["threshold_party_pct"],
        "threshold_alliance_pct": CFG["threshold_alliance_pct"],
        "registered": nat_reg,
        "ballots_cast": nat_cast,
        "valid": nat_valid,
        "invalid": nat_invalid,
        "turnout_pct": round(100 * nat_cast / nat_reg, 2),
        "total_seats": CFG["total_seats"],
        "stations": int(len(results)),
        "parties": national_parties,
    }
    (DATA / "national.json").write_text(
        json.dumps(national, ensure_ascii=False, indent=2), encoding="utf-8")

    # --- per-marz JSON (winner, margin, shares, turnout) ---
    marz_json = {}
    party_marz_pct = {pid: {} for pid in PARTY_IDS}
    for _, row in marz.iterrows():
        sh = shares_dict(row)
        info = winner_info(sh)
        iso = row["marz_iso"]
        marz_json[iso] = {
            "iso": iso,
            "name_en": row["marz_en"],
            "name_hy": MARZ_INV[row["marz_en"]][3],
            "name_fr": MARZ_INV[row["marz_en"]][2],
            "registered": int(row["registered"]),
            "ballots_cast": int(row["ballots_cast"]),
            "valid": int(row["valid"]),
            "invalid": int(row["invalid"]),
            "turnout_pct": float(row["turnout_pct"]),
            "stations": int(row["stations"]),
            "winner": info["winner"],
            "winner_pct": info["winner_pct"],
            "runner_up": info["runner_up"],
            "margin": info["margin"],
            "order": info["order"],
            "shares": sh,
        }
        for pid in PARTY_IDS:
            party_marz_pct[pid][iso] = sh[pid]["pct"]
    if len(marz_json) != len(MARZ):
        sys.exit(f"Integrity error: produced {len(marz_json)} provinces, expected "
                 f"{len(MARZ)} — a non-province (diaspora/electronic/abroad?) row "
                 f"may have leaked in. ISOs: {sorted(marz_json)}")
    marz_valid_sum = sum(m["valid"] for m in marz_json.values())
    if marz_valid_sum + out_of_marz_votes != nat_valid:
        sys.exit(f"Integrity error: marz valid ({marz_valid_sum}) + out-of-marz "
                 f"({out_of_marz_votes}) != national valid ({nat_valid}).")
    (DATA / "marz.json").write_text(
        json.dumps(marz_json, ensure_ascii=False, indent=2), encoding="utf-8")

    # --- per-party JSON (national + per-marz pct + best/worst marz) ---
    parties_out = []
    for p in national_parties:
        pid = p["id"]
        pm = party_marz_pct[pid]
        ranked = sorted(pm.items(), key=lambda kv: kv[1], reverse=True)
        parties_out.append({
            **p,
            "by_marz": pm,
            "best_marz": ranked[0][0] if ranked else None,
            "worst_marz": ranked[-1][0] if ranked else None,
            "won_marz": [iso for iso, m in marz_json.items() if m["winner"] == pid],
        })
    (DATA / "parties.json").write_text(
        json.dumps(parties_out, ensure_ascii=False, indent=2), encoding="utf-8")

    # --- meta / provenance + integrity checks ---
    meta = {
        "generated_utc": _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        "source": SOURCE,
        "marzer": {iso: {"en": en, "fr": fr, "hy": hy}
                   for hy_, (iso, en, fr, hy) in MARZ.items()},
        "integrity": {
            "stations_parsed": int(len(results)),
            "provinces": len(marz_json),
            "national_valid_votes": nat_valid,
            "national_ballots_cast": nat_cast,
            "out_of_province_stations": int(len(unmatched)),
            "out_of_province_valid_votes": out_of_marz_votes,
            "seat_winner_votes": {pid: int(results[pid].sum()) for pid in CFG["seats"]},
        },
        "files": {
            "stations": [f"data/{ELECTION}/clean/stations.parquet", f"data/{ELECTION}/clean/stations.csv"],
            "communities": [f"data/{ELECTION}/clean/communities.parquet", f"data/{ELECTION}/clean/communities.csv"],
            "settlements": [f"data/{ELECTION}/clean/settlements.parquet", f"data/{ELECTION}/clean/settlements.csv"],
            "marz": [f"data/{ELECTION}/clean/marz.parquet", f"data/{ELECTION}/clean/marz.csv"],
        },
        "levels": {
            "communities": int(comm["community_hy"].nunique()),
            "settlements": int(sett["locality_hy"].nunique()),
        },
    }
    (DATA / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    # --- console summary ---
    print("OK — parsed", len(results), "polling stations")
    print(f"   national turnout : {national['turnout_pct']}%  "
          f"({nat_cast:,} / {nat_reg:,})")
    for p in national_parties[:5]:
        print(f"   {p['name_en']:<28} {p['votes']:>9,}  {p['pct']:>6.2f}%  "
              f"{p['seats']} seats")
    print("   marz winners:", {iso: m["winner"] for iso, m in marz_json.items()})


if __name__ == "__main__":
    build()
